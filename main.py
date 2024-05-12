from fastapi import FastAPI, File, UploadFile
from fastapi.responses import JSONResponse
from openpyxl import load_workbook
from fastapi import FastAPI
import pandas as pd
import tkinter as tk
from pathlib import Path
import os
import re
import math
from datetime import datetime, time, timedelta
from tkinter import Tk, filedialog
from openpyxl import load_workbook
app = FastAPI()
@app.get("/")
async def root():

    return {"message": "Hello World"}


@app.get("/hello/{name}")
async def say_hello(name: str):
    return {"message": f"Hello {name}"}









penalizacion=timedelta()
progress=-1
pena=False
dp12=0
matrizPenalizacion=[["WP#","Coor (Lat)","Coor (Lon)","CoorUS (Lat)","CoorUS (Lon)","Causa","Valor Esperado","Valor Usuario","Penalizacion","Penalizacion Total"]]
row=[]
radio=0



# Función para convertir coordenadas DMS a DD
def dms_to_dd(col):
    list=[]
    for i in col:

        c=0
        grad=""
        mins=""
        pos=""
        for j in i:

            if (j=="°"):
                c=1
            if (j==","):
                c=2
            if (c==0):
                grad+=j
            if(c==1):
                mins+=j
            if(c==2):
                pos+=j
        grad = float(grad)
        mins = float(mins[1:])
        pos= pos[1:]
        if (pos in ("N", "E")):
            result = grad + (mins / 60)
        else:
            result = (grad + (mins / 60)) * -1
        list.append(result)
    df = pd.DataFrame(list, columns=['latitude'])
    return df


def carga():
    try:
        # Crear una ventana Tkinter sin mostrarla
        root = Tk()
        root.withdraw()

        # Mostrar el diálogo de selección de archivo y obtener la ruta del archivo seleccionado
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])

        if file_path:
            # Cargar el archivo XLSX seleccionado
            wb = load_workbook(filename=file_path)
            sheet = wb.active
            data = sheet.values
            columns = next(data)
            df = pd.DataFrame(data, columns=columns)
            df = df.dropna(how='all')
            df = df.dropna(axis=1, how='all')

            return df
        else:
            print("No se seleccionó ningún archivo.")

    except Exception as e:
        print(f"No se pudo cargar el archivo: {e}")


# capturo un archivo de forma "busqueda ubicacion" y creo ahi la matriz de penalizaciones
def gpxUsuario():
    root = tk.Tk()
    root.withdraw()
    # Mostrar el diálogo para seleccionar el archivo
    ruta_seleccionada = filedialog.askopenfilename()
    # Convertir la ruta seleccionada a un objeto Path
    ruta_absoluta = Path(ruta_seleccionada)

    # Verificar si se seleccionó un archivo
    if ruta_absoluta.is_file():
        # El archivo seleccionado existe
        with open(ruta_absoluta, "r") as archivo:
            contenido = archivo.read()

            # Buscar la primera ocurrencia de "lat=" en el contenido
            indice_inicio = contenido.find("lat=")

            if indice_inicio != -1:
                # Se encontró "lat=", obtener el contenido a partir de esa posición
                inicio = contenido[indice_inicio:]


                # Patrones para encontrar latitud, longitud y tiempo en la cadena de texto
                patron_latitud = r'lat="([^"]+)"'
                patron_longitud = r'lon="([^"]+)"'
                patron_tiempo = r'<time>([^<]+)</time>'

                # Encontrar todas las coincidencias de latitud, longitud y tiempo en la cadena de texto
                coincidencias_latitud = re.findall(patron_latitud, inicio)
                coincidencias_longitud = re.findall(patron_longitud, inicio)
                coincidencias_tiempo = re.findall(patron_tiempo, inicio)

                # Combinar las coincidencias en una matriz
                matriz_resultante = list(zip(coincidencias_latitud, coincidencias_longitud, coincidencias_tiempo))
                df = pd.DataFrame(matriz_resultante, columns=['lat', 'lon', 'time'])
                nombre_archivo = os.path.basename(ruta_seleccionada)
                nombre_archivo = nombre_archivo[:-3]
                print(f"Nombre del archivo seleccionado: {nombre_archivo}")


                return (df)

            else:
                print("No se encontró 'lat=' en el archivo.")


    else:
        print("No se seleccionó ningún archivo.")

def segToHours(time):
    horas = int(time // 3600)  # Obtener la parte entera de la división
    minutos = int((time % 3600) // 60)  # Obtener los minutos
    segundos = int(time % 60)  # Obtener los segundos restantes

    # Imprimir el resultado
    print("Tiempo en horas, minutos y segundos:")
    print(f"{horas} horas, {minutos} minutos, {segundos} segundos")
def totalTime(time):
    # Strings de fecha y hora
    fecha_hora_str1 = time[0]
    fecha_hora_str2 = time.iloc[-1]

    # Convertir los strings a objetos de fecha y hora
    fecha_hora1 = datetime.fromisoformat(fecha_hora_str1.replace('Z', '+00:00'))
    fecha_hora2 = datetime.fromisoformat(fecha_hora_str2.replace('Z', '+00:00'))

    # Calcular la diferencia de tiempo
    diferencia = fecha_hora2 - fecha_hora1

    # Imprimir la diferencia de tiempo en segundos
    #print("Diferencia de tiempo en segundos:", diferencia.total_seconds())
    segToHours(diferencia.total_seconds())

def totalTime2(time1,time2):
    # Strings de fecha y hora
    fecha_hora_str1 = time1
    fecha_hora_str2 = time2

    # Convertir los strings a objetos de fecha y hora
    fecha_hora1 = datetime.fromisoformat(fecha_hora_str1.replace('Z', '+00:00'))
    fecha_hora2 = datetime.fromisoformat(fecha_hora_str2.replace('Z', '+00:00'))

    # Calcular la diferencia de tiempo
    diferencia = fecha_hora2 - fecha_hora1

    # Imprimir la diferencia de tiempo en segundos
    #print("Diferencia de tiempo en segundos:", diferencia.total_seconds())
    return (diferencia.total_seconds())

def matchEx(tupla,dataUser):
    global progress
    for i, fila2 in dataUser.iterrows():
        if i < progress:
            continue
        tupla2 = (fila2['lat'], fila2['lon'])
        redondeado_tupla = tuple(round(float(coord), 3) for coord in tupla)
        redondeado_tupla2 = tuple(round(float(coord), 3) for coord in tupla2)


        if (redondeado_tupla == redondeado_tupla2 and (i-progress)<len(dataUser)/20):
            progress=i
            return (True,fila2)


    return (False,fila2)

def neutral(tupla,dataUser,radio,tiempoN):

    global penalizacion, progress
    mach=True
    while mach:
        c = 0
        for i, fila2 in dataUser.iterrows():
            if i <= progress:
                c += 1
                time1=fila2['time']
                continue

            distancia_metros = distancia_euclidiana(tupla[0], tupla[1], fila2['lat'], fila2['lon'])
            rango = float(radio / 1000)
            if (distancia_metros>rango):
                time2=fila2['time']
                timeT=totalTime2(time1,time2)
                timeT=int(timeT/60)
                #print(f"Tiempo de Espera: {timeT}/ Tiempo obligatorio: {tiempoN}")
                timeT = math.floor(timeT)
                mach=False
                progress=i

                break
    if(timeT<tiempoN-1):
        t=tiempoN-timeT
        t=math.floor(t*2)

        penalizacion += timedelta(minutes=t)
        ###############matrizPenalizacion.append([])
        #print("PENALIZACION POR NO NEUTRALIZACION------->",t)
        #print("PENALIZACION TOTAL:  ",penalizacion)
        row.extend(["PENALIZACION NO NEUTRALIZACION", f"00:{tiempoN}:00", f"00:{timeT}:00", f"00:{t}:00", penalizacion])
        matrizPenalizacion.append(row)

def match(tupla,dataUser,radio1,tiempo_str):
    long=len(dataUser)
    matchExatle=matchEx(tupla,dataUser)
    mach=matchExatle[0]
    global penalizacion,pena,progress,row,radio
    pena=False
    while not mach:
        c=0
        min=500
        for i, fila2 in dataUser.iterrows():
            if i<=progress:
                c+=1
                continue
            distancia_metros = distancia_euclidiana(tupla[0], tupla[1], fila2['lat'], fila2['lon'])
            rango = float(radio1)#111320 porque 1 grado de longitud en la superficie de la Tierra equivale a aproximadamente 111320 metros
            if(radio1>radio1+75):
                radio=None
                return pd.DataFrame() #No hubo match
            if (distancia_metros < rango and progress<i ):
                if (distancia_metros<= min):
                    min=distancia_metros
                    matchExatle = (True, fila2)
                    progress=i

            elif (distancia_metros > rango and matchExatle[0]):
                mach = True
                if (pena and tiempo_str is not None):

                    penalizacion += timedelta(seconds=tiempo_str.second)
                    penalizacion += timedelta(minutes=tiempo_str.minute)
                    penalizacion += timedelta(hours=tiempo_str.hour)
                    #print("PENALIZACION TOTAL: ", penalizacion)

                break

            c+=1
            if (i==long-1):
                if(min<1):
                    mach = True
                    if (pena and tiempo_str is not None):
                        penalizacion += timedelta(seconds=tiempo_str.second)
                        penalizacion += timedelta(minutes=tiempo_str.minute)
                        penalizacion += timedelta(hours=tiempo_str.hour)
                        # print("PENALIZACION TOTAL: ", penalizacion)
                    break

                #print("no match con ",radio)
                else:
                    radio1+=25 #aumento el radio 10 metros para que lo capture
                    radio=radio1
                    mach=False
                    pena=True

    return (matchExatle[1])


def matchWp(dataFWP,dataUser):

    # Iterar sobre cada fila de la tabla más pequeña y comparar con cada fila de la tabla más grande
    global penalizacion,pena,row,radio

    for index,fila1 in dataFWP.iterrows():
        #print(fila1)
        #print(f"WP {index} actual:", punto1)
        row = []

        punto1 = (fila1['latitude'], fila1['longitude'])

        if(fila1['type']=='NaN' or fila1['type']==None ):
            continue
        #print(f"WP {fila1['WP #']}:", punto1)

        row.extend([fila1['WP #'],round(fila1['latitude'], 3), round(fila1['longitude'], 3)])

        MMU= match(punto1, dataUser, fila1['ratius'], fila1['penalization'])

        speedProm=0
        if not (MMU.empty):
            punto2=(MMU['lat'],MMU['lon'])
            #print(f"CU {progress}:", punto2)
            row.extend([round(float(MMU['lat']), 3), round(float(MMU['lon']), 3)])


            if(fila1['type']=='N'):
                neutral(punto1,dataUser,fila1['ratius'], int(fila1['NeutralizacionT'].minute))
            if (fila1['type']=='DZ'):
                filaanterior=fila1['distance']
                gti=MMU['time']
            elif (fila1['type']=='FZ'):
                distancia=fila1['distance']-filaanterior
                gtf = MMU['time']

                speedProm=speed(distancia,gti,gtf)
                #print("VELOCIDAD PROMEDIO: ",speedProm)

            if pena:
                #print("PENALIZACION NO MATCH------->", fila1['penalization'])
                #print("PENALIZACION TOTAL: ", penalizacion)
                row.extend(["PENALIZACION NO MATCH",str(int(fila1['ratius']))+" Mts",str(int(radio))+" Mts",fila1['penalization'],penalizacion])
                matrizPenalizacion.append(row)


            if(speedProm>=float(fila1['speed'])+1 and not pena):        #si no se penaliza el NO MATCH
                t=speedProm-float(fila1['speed'])
                p=math.floor(t/10)
                t+=(p*10)
                t=math.floor(t)
                #print("PENALIZACION SPEED------->",t)
                penalizacion += timedelta(minutes=t)

                #print("PENALIZACION TOTAL: ", penalizacion)
                row.extend(["PENALIZACION VELOCIDAD", str(int(fila1['speed']))+" Km/h", str(int(speedProm))+" Km/h", f"00:{t}:00", penalizacion])
                matrizPenalizacion.append(row)



        else:
            #print("PENALIZACION NO MATCH------->", fila1['penalization'])
            #print(f" Se alejo un radio de mas de 100 metros, del origen del punto")
            if fila1['penalization'] is not None:
                penalizacion += timedelta(seconds=fila1['penalization'].second)
                penalizacion += timedelta(minutes=fila1['penalization'].minute)
                penalizacion += timedelta(hours=fila1['penalization'].hour)
            #print("PENALIZACION TOTAL: ", penalizacion)
            row.extend(["---","---","PENALIZACION NO MATCH", str(int(fila1['ratius']))+" Mts", None, fila1['penalization'], penalizacion])
            matrizPenalizacion.append(row)

        #print("\n")#df_coincidencias = pd.DataFrame(coincidencias_exactas, columns=['id_tabla1', 'id_tabla2', 'distancia'])



def speed(distancia,gti,gtf):

    tiempo=totalTime2(gti,gtf)
    tiempo=tiempo/3600
    return (distancia/tiempo)

def distancia_euclidiana(x1, y1, x2, y2):
    return math.sqrt(((float(x2) - float(x1))**2 + (float(y2) - float(y1))**2) * 111320)



start= datetime.now()

#pd.set_option('display.max_rows', None)
#pd.set_option('display.max_columns', 10)

#print(matrizCompleta)
matrizCompleta = carga()
matrizCompleta['latitude'] = dms_to_dd(matrizCompleta['latitude'])
matrizCompleta['longitude'] = dms_to_dd(matrizCompleta['longitude'])



gpxCorredor=gpxUsuario()

totalTime(gpxCorredor['time'])


matchWp(matrizCompleta,gpxCorredor)

df_penalizacion = pd.DataFrame(matrizPenalizacion[1:], columns=matrizPenalizacion[0])

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', 10)

print(df_penalizacion)
end=datetime.now()

print("\n \nTIEMPO DE EJECUCION =",end-start)

